
import requests
import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# ── EXTRACCIÓN ──
url = "https://api.open-meteo.com/v1/forecast"
parametros = {
    "latitude": 19.4326,
    "longitude": -99.1332,
    "daily": ["temperature_2m_max", "temperature_2m_min", "precipitation_sum", "windspeed_10m_max"],
    "timezone": "America/Mexico_City",
    "past_days": 30
}
response = requests.get(url, params=parametros)
data = response.json()

# ── TRANSFORMACIÓN ──
daily_data = data["daily"]
df = pd.DataFrame({
    "date": daily_data["time"],
    "temp_max": daily_data["temperature_2m_max"],
    "temp_min": daily_data["temperature_2m_min"],
    "precipitation": daily_data["precipitation_sum"],
    "wind_max": daily_data["windspeed_10m_max"]
})
df["date"] = pd.to_datetime(df["date"])
df["avg_temp"] = (df["temp_max"] + df["temp_min"]) / 2
df["range_temp"] = df["temp_max"] - df["temp_min"]
df["rain_day"] = (df["precipitation"] > 0).astype(int)

# ── CARGA A SQLITE ──
conn = sqlite3.connect(r"C:\Users\PC\OneDrive\Documentos\Proyecto Ventas\clima_cdmx.db")
df.to_sql("clima_diario", conn, if_exists="replace", index=False)

# ── REPORTE EXCEL ──
wb = Workbook()
wb.remove(wb.active)

color_header = "1a6fa8"
color_subheader = "5bc0eb"

def estilo_header(celda, color):
    celda.font = Font(bold=True, color="FFFFFF", size=12)
    celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    celda.alignment = Alignment(horizontal="center", vertical="center")

def escribir_hoja(wb, nombre_hoja, df, titulo):
    ws = wb.create_sheet(title=nombre_hoja)
    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    estilo_header(ws["A1"], color_header)
    ws.row_dimensions[1].height = 30
    for col, nombre in enumerate(df.columns, start=1):
        celda = ws.cell(row=2, column=col, value=nombre)
        estilo_header(celda, color_subheader)
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        col_letter = ws.cell(row=2, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max_length + 4
    return ws

q1 = pd.read_sql("""
    SELECT COUNT(*) as total_days,
           ROUND(AVG(temp_max), 1) AS avg_temp_max,
           ROUND(AVG(temp_min), 1) AS avg_temp_min,
           ROUND(AVG(avg_temp), 1) AS avg_temp_general,
           ROUND(SUM(precipitation), 1) AS precipitation_total,
           SUM(rain_day) AS raining_days
    FROM clima_diario
""", conn)

q2 = pd.read_sql("""
    SELECT date, temp_max, temp_min, avg_temp, precipitation
    FROM clima_diario
    ORDER BY temp_max DESC
    LIMIT 5
""", conn)

q3 = pd.read_sql("""
    SELECT date, precipitation, temp_max, wind_max
    FROM clima_diario
    WHERE rain_day = 1
    ORDER BY precipitation DESC
    LIMIT 5
""", conn)

escribir_hoja(wb, "General Resume", q1, "General Resume — CDMX Weather")
escribir_hoja(wb, "Hottest Days", q2, "Top 5 Hottest Days")
escribir_hoja(wb, "Rainiest Days", q3, "Top 5 Rainiest Days")
escribir_hoja(wb, "Raw Data", df, "Daily Weather Data")

fecha = datetime.now().strftime("%Y%m%d_%H%M")
ruta = rf"C:\Users\PC\OneDrive\Documentos\Proyecto Ventas\reports\reporte_clima_{fecha}.xlsx"
wb.save(ruta)
print(f"Reporte generado: {ruta}")
